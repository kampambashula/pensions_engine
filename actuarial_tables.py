"""
actuarial_tables.py
--------------------
Actuarial table loader and lookup engine.

Sources:
  a55_mortality_tables.xlsx   — a(55) immediate annuitant mortality (male / female)
                                One-year select rates: Duration 0 and Durations 1+
  Ill_health_rates_cmiwp50ac04rates-v2.xls
                              — CMI WP50 critical illness / ill-health rates
                                ACMNL04 (male non-smoker) and ACFNL04 (female non-smoker)
                                q-type rates by age exact and duration 0-1-2-3-4-5+

Provides:
  - MortalityTable       — q[x] lookup for a(55) m/f tables
  - IllHealthTable       — q[x] lookup for ACMNL04/ACFNL04 by duration
  - annuity_factor()     — whole-life annuity computed from a(55) mortality
  - get_mortality_qx()   — unified public lookup used by calculation_engine.py
  - get_ill_health_qx()  — unified public lookup used by calculation_engine.py

Author: Pension Actuarial Systems — Actuarial Model Layer
"""

from __future__ import annotations

import logging
from functools import lru_cache
from pathlib import Path
from typing import Literal

import numpy as np
import pandas as pd

log = logging.getLogger("actuarial_tables")

BASE_DIR    = Path(__file__).resolve().parent
MORTALITY_FILE  = BASE_DIR / "data" / "a55_mortality_tables.xlsx"
ILL_HEALTH_FILE = BASE_DIR / "data" / "ill_health_rates.xls"

# Fallback paths: look in the uploads directory if data/ copies aren't present
_UPLOAD_DIR = Path("/mnt/user-data/uploads")
_MORT_UPLOAD  = _UPLOAD_DIR / "a55_mortality_tables__1_.xlsx"
_IH_UPLOAD    = _UPLOAD_DIR / "Ill_health_rates_cmiwp50ac04rates-v2__1_.xls"

Gender = Literal["Male", "Female", "Unknown"]


# ---------------------------------------------------------------------------
# a(55) Mortality Table
# ---------------------------------------------------------------------------

class MortalityTable:
    """
    a(55) immediate annuitant mortality table.

    Stores q[x] (one-year select death probabilities) for:
      - Duration 0  : first policy year
      - Durations 1+: all subsequent years (ultimate)

    Usage:
        table = MortalityTable.load()
        qx = table.qx(age=60, gender="Male", duration=1)
    """

    def __init__(self, male_df: pd.DataFrame, female_df: pd.DataFrame):
        # index by integer age
        self._male   = male_df.set_index("age")
        self._female = female_df.set_index("age")

    @classmethod
    def load(cls) -> "MortalityTable":
        path = MORTALITY_FILE if MORTALITY_FILE.exists() else _MORT_UPLOAD
        if not path.exists():
            log.warning("Mortality table file not found — using default rates.")
            return cls._build_default()

        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True)

        def _read_sheet(ws) -> pd.DataFrame:
            rows = []
            for row in ws.iter_rows(values_only=True):
                age = row[0]
                if age is None or not isinstance(age, (int, float)):
                    continue
                age = int(age)
                dur0 = row[1] if row[1] is not None else np.nan
                dur1plus = row[2] if row[2] is not None else np.nan
                # For ages where duration 0 is missing (older ages), use dur1+
                if pd.isna(dur0) and not pd.isna(dur1plus):
                    dur0 = dur1plus
                rows.append({"age": age, "dur0": float(dur0) if not pd.isna(dur0) else 1.0,
                             "dur1plus": float(dur1plus) if not pd.isna(dur1plus) else 1.0})
            return pd.DataFrame(rows)

        male_df   = _read_sheet(wb["a(55)m"])
        female_df = _read_sheet(wb["a(55)f"])
        log.info("Mortality tables loaded: a(55)m ages %d-%d, a(55)f ages %d-%d",
                 male_df.age.min(), male_df.age.max(),
                 female_df.age.min(), female_df.age.max())
        return cls(male_df, female_df)

    def qx(self, age: float, gender: Gender, duration: int = 1) -> float:
        """
        Return q[x] for given age, gender and duration.
        duration=0  → first year select rate
        duration>=1 → ultimate rate (Durations 1+)
        """
        tbl = self._male if gender != "Female" else self._female
        age_i = max(int(age), tbl.index.min())
        age_i = min(age_i, tbl.index.max())

        try:
            row = tbl.loc[age_i]
            return float(row["dur0"] if duration == 0 else row["dur1plus"])
        except (KeyError, TypeError):
            return 1.0   # terminal age — assume certain death

    def annuity_factor(
        self,
        age: float,
        gender: Gender,
        interest_rate: float = 0.08,
        max_age: int = 110,
    ) -> float:
        """
        Compute a whole-life annuity factor ä_x at given interest rate
        using a(55) mortality.

          ä_x = Σ_{t=0}^{ω-x} (t_p_x) × v^t
          where v = 1/(1+i)  and  t_p_x = Π_{k=0}^{t-1} (1 - q_{x+k})

        Returns the annuity factor (number of annual payments of 1 unit).
        """
        v = 1.0 / (1.0 + interest_rate)
        tpx   = 1.0
        annuity = 0.0

        for t in range(max_age - int(age) + 1):
            annuity += tpx * (v ** t)
            qx_t    = self.qx(age + t, gender, duration=1 if t > 0 else 0)
            tpx    *= (1.0 - qx_t)
            if tpx < 1e-10:
                break

        return annuity

    def life_expectancy(self, age: float, gender: Gender, max_age: int = 110) -> float:
        """
        Curtate future lifetime e_x = Σ t_p_x  (t from 1 to omega-x)
        """
        tpx = 1.0
        ex  = 0.0
        for t in range(1, max_age - int(age) + 1):
            qx_t = self.qx(age + t - 1, gender, duration=1)
            tpx *= (1.0 - qx_t)
            ex  += tpx
            if tpx < 1e-10:
                break
        return ex

    @classmethod
    def _build_default(cls) -> "MortalityTable":
        """Fallback: simple Makeham-type rates when file unavailable."""
        ages  = range(18, 117)
        # Very simple approximation: q_x = A + B*c^x
        A, B, c_m = 0.0007, 0.00005, 1.09
        A_f, B_f, c_f = 0.0005, 0.00003, 1.09
        male_rows   = [{"age": x, "dur0": min(A + B * c_m**x, 1.0), "dur1plus": min(A + B * c_m**x * 1.2, 1.0)} for x in ages]
        female_rows = [{"age": x, "dur0": min(A_f + B_f * c_f**x, 1.0), "dur1plus": min(A_f + B_f * c_f**x * 1.2, 1.0)} for x in ages]
        return cls(pd.DataFrame(male_rows), pd.DataFrame(female_rows))


# ---------------------------------------------------------------------------
# Ill-Health (Critical Illness) Table  — CMI WP50 ACMNL04 / ACFNL04
# ---------------------------------------------------------------------------

class IllHealthTable:
    """
    CMI WP50 critical illness / ill-health inception rates.

    ACMNL04 — Male non-smoker
    ACFNL04 — Female non-smoker

    Rates are q-type annual claim rates by age exact and duration
    (0, 1, 2, 3, 4, 5+).

    Usage:
        table = IllHealthTable.load()
        qx = table.qx(age=45, gender="Male", duration=5)
    """

    def __init__(self, male_df: pd.DataFrame, female_df: pd.DataFrame):
        self._male   = male_df.set_index("age")
        self._female = female_df.set_index("age")

    @classmethod
    def load(cls) -> "IllHealthTable":
        path = ILL_HEALTH_FILE if ILL_HEALTH_FILE.exists() else _IH_UPLOAD
        if not path.exists():
            log.warning("Ill-health table file not found — using default rates.")
            return cls._build_default()

        def _read_sheet(xls, sheet_name: str) -> pd.DataFrame:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine="xlrd")
            rows = []
            for _, row in df.iterrows():
                age = row.iloc[0]
                if age is None or pd.isna(age):
                    continue
                try:
                    age = int(age)
                except (ValueError, TypeError):
                    continue
                # q-type columns: dur0=col1, dur1=col2, dur2=col3, dur3=col4, dur4=col5, dur5+=col6
                def _safe(v):
                    return float(v) if not pd.isna(v) else np.nan
                rows.append({
                    "age":   age,
                    "dur0":  _safe(row.iloc[1]),
                    "dur1":  _safe(row.iloc[2]),
                    "dur2":  _safe(row.iloc[3]),
                    "dur3":  _safe(row.iloc[4]),
                    "dur4":  _safe(row.iloc[5]),
                    "dur5p": _safe(row.iloc[6]),
                })
            df_out = pd.DataFrame(rows)
            # Forward-fill dur0 from dur5p where missing (high ages)
            for col in ["dur0","dur1","dur2","dur3","dur4"]:
                df_out[col] = df_out[col].fillna(df_out["dur5p"])
            return df_out

        xls = pd.ExcelFile(str(path), engine="xlrd")
        male_df   = _read_sheet(xls, "ACMNL04")
        female_df = _read_sheet(xls, "ACFNL04")
        log.info("Ill-health tables loaded: ACMNL04 ages %d-%d, ACFNL04 ages %d-%d",
                 male_df.age.min(), male_df.age.max(),
                 female_df.age.min(), female_df.age.max())
        return cls(male_df, female_df)

    def qx(self, age: float, gender: Gender, duration: int = 5) -> float:
        """
        Return ill-health inception rate for age / gender / duration.
        duration: 0,1,2,3,4 → policy year; 5+ = ultimate
        """
        tbl   = self._male if gender != "Female" else self._female
        age_i = max(int(age), tbl.index.min())
        age_i = min(age_i, tbl.index.max())
        dur_col = {0:"dur0", 1:"dur1", 2:"dur2", 3:"dur3", 4:"dur4"}.get(duration, "dur5p")
        try:
            return float(tbl.loc[age_i, dur_col])
        except (KeyError, TypeError):
            return 0.0

    @classmethod
    def _build_default(cls) -> "IllHealthTable":
        ages = range(18, 111)
        male_rows   = [{"age": x, "dur0": 0.0005 + x * 0.00008, "dur1": 0.0007 + x * 0.0001,
                        "dur2": 0.0007 + x * 0.0001, "dur3": 0.0007 + x * 0.0001,
                        "dur4": 0.0007 + x * 0.0001, "dur5p": 0.001 + x * 0.00012} for x in ages]
        female_rows = [{"age": x, "dur0": 0.0003 + x * 0.00005, "dur1": 0.0005 + x * 0.00007,
                        "dur2": 0.0005 + x * 0.00007, "dur3": 0.0005 + x * 0.00007,
                        "dur4": 0.0005 + x * 0.00007, "dur5p": 0.0007 + x * 0.00009} for x in ages]
        return cls(pd.DataFrame(male_rows), pd.DataFrame(female_rows))


# ---------------------------------------------------------------------------
# Module-level singletons (load once, reuse everywhere)
# ---------------------------------------------------------------------------

@lru_cache(maxsize=1)
def get_mortality_table() -> MortalityTable:
    return MortalityTable.load()


@lru_cache(maxsize=1)
def get_ill_health_table() -> IllHealthTable:
    return IllHealthTable.load()


# ---------------------------------------------------------------------------
# Public lookup helpers (used by calculation_engine.py)
# ---------------------------------------------------------------------------

def get_mortality_qx(age: float, gender: Gender, duration: int = 1) -> float:
    """q[x] from a(55) table."""
    return get_mortality_table().qx(age, gender, duration)


def get_ill_health_qx(age: float, gender: Gender, duration: int = 5) -> float:
    """Ill-health inception rate from CMI WP50 table."""
    return get_ill_health_table().qx(age, gender, duration)


def annuity_factor(age: float, gender: Gender, interest_rate: float = 0.08) -> float:
    """Whole-life annuity factor ä_x from a(55) table at given interest rate."""
    return get_mortality_table().annuity_factor(age, gender, interest_rate)


def life_expectancy(age: float, gender: Gender) -> float:
    """Curtate future lifetime e_x from a(55) table."""
    return get_mortality_table().life_expectancy(age, gender)


# ---------------------------------------------------------------------------
# Table summary helpers (for dashboard display)
# ---------------------------------------------------------------------------

def mortality_table_summary() -> pd.DataFrame:
    """Return a display-ready mortality table for both genders, ages 20-100."""
    t = get_mortality_table()
    rows = []
    for age in range(20, 101, 5):
        rows.append({
            "Age": age,
            "Male q[x] Dur 0":   f"{t.qx(age, 'Male',   0):.5f}",
            "Male q[x] Dur 1+":  f"{t.qx(age, 'Male',   1):.5f}",
            "Female q[x] Dur 0": f"{t.qx(age, 'Female', 0):.5f}",
            "Female q[x] Dur 1+":f"{t.qx(age, 'Female', 1):.5f}",
            "Male ä_x (8%)":     f"{t.annuity_factor(age, 'Male',   0.08):.3f}",
            "Female ä_x (8%)":   f"{t.annuity_factor(age, 'Female', 0.08):.3f}",
        })
    return pd.DataFrame(rows)


def ill_health_table_summary() -> pd.DataFrame:
    """Return a display-ready ill-health table for both genders, ages 20-65."""
    t = get_ill_health_table()
    rows = []
    for age in range(20, 66, 5):
        rows.append({
            "Age": age,
            "Male Dur 0":  f"{t.qx(age, 'Male',   0):.5f}",
            "Male Dur 1":  f"{t.qx(age, 'Male',   1):.5f}",
            "Male Dur 5+": f"{t.qx(age, 'Male',   5):.5f}",
            "Female Dur 0":f"{t.qx(age, 'Female', 0):.5f}",
            "Female Dur 1":f"{t.qx(age, 'Female', 1):.5f}",
            "Female Dur 5+":f"{t.qx(age, 'Female',5):.5f}",
        })
    return pd.DataFrame(rows)


def annuity_factor_grid() -> pd.DataFrame:
    """Return annuity factors at various ages and interest rates."""
    t = get_mortality_table()
    rates = [0.05, 0.06, 0.07, 0.08, 0.09, 0.10, 0.12]
    rows = []
    for age in range(20, 75, 5):
        row = {"Age": age}
        for r in rates:
            row[f"Male {r:.0%}"]   = round(t.annuity_factor(age, "Male",   r), 3)
            row[f"Female {r:.0%}"] = round(t.annuity_factor(age, "Female", r), 3)
        rows.append(row)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# CLI test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=== a(55) Mortality Table — Sample ===")
    t = get_mortality_table()
    for age in [30, 40, 50, 55, 60, 65]:
        qm  = t.qx(age, "Male")
        qf  = t.qx(age, "Female")
        am  = t.annuity_factor(age, "Male",   0.08)
        af  = t.annuity_factor(age, "Female", 0.08)
        lem = t.life_expectancy(age, "Male")
        lef = t.life_expectancy(age, "Female")
        print(f"  Age {age:3d} | qx M={qm:.5f} F={qf:.5f} | "
              f"ä_x(8%) M={am:.3f} F={af:.3f} | "
              f"e_x M={lem:.1f} F={lef:.1f}")

    print("\n=== CMI WP50 Ill-Health Table — Sample ===")
    ih = get_ill_health_table()
    for age in [30, 40, 50, 60]:
        qm  = ih.qx(age, "Male",   5)
        qf  = ih.qx(age, "Female", 5)
        print(f"  Age {age:3d} | q[x] Dur5+ M={qm:.5f}  F={qf:.5f}")
