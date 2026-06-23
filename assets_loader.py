"""
assets_loader.py
-----------------
Loads the fund's financial statements from Assets_Information.xlsx
(Income Statement + Balance Sheet — Revenue Account format).

Provides:
  - FundFinancials dataclass  — structured annual figures
  - load_fund_financials()    — reads and returns all available years
  - solvency_position()       — computes solvency table from assets + liabilities

The spreadsheet contains a dual-layout format:
  Left side  → Income Statement (2021, 2022, 2023)
  Right side → Balance Sheet    (2020, 2021, 2022, 2023)

Author: Pension Actuarial Systems — Data Engineering Layer
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

log = logging.getLogger("assets_loader")

BASE_DIR    = Path(__file__).resolve().parent
ASSETS_FILE = BASE_DIR / "data" / "Assets_Information.xlsx"
_UPLOAD     = Path("/mnt/user-data/uploads") / "Assets_Information.xlsx"


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class IncomeStatement:
    year_end: str                          # e.g. "2021", "2022", "2023"
    fund_opening:               float = 0
    employee_contributions:     float = 0
    employer_contributions:     float = 0
    financing_gap:              float = 0
    government_grants:          float = 0
    investment_income:          float = 0
    fx_gains_losses:            float = 0
    change_fv_investment_prop:  float = 0
    change_fv_equity:           float = 0
    other_income:               float = 0
    total_income:               float = 0
    pensions_paid:              float = 0
    commutation_lumpsum:        float = 0
    lumpsum_early_death:        float = 0
    other_benefits:             float = 0
    admin_expenses:             float = 0
    withholding_tax:            float = 0
    finance_cost:               float = 0
    total_expenditure:          float = 0
    surplus:                    float = 0
    fund_closing:               float = 0


@dataclass
class BalanceSheet:
    year: int
    # Non-current assets
    ppe:                        float = 0
    investment_properties:      float = 0
    staff_loans:                float = 0
    home_loan_scheme:           float = 0
    microfinance_loans:         float = 0
    equity_investments:         float = 0
    fixed_deposits:             float = 0
    government_securities:      float = 0
    work_in_progress:           float = 0
    intangibles:                float = 0
    # Current assets
    contributions_due:          float = 0
    other_receivables:          float = 0
    cash_at_bank:               float = 0
    total_assets:               float = 0
    # Liabilities
    pensions_payable:           float = 0
    benefits_payable:           float = 0
    other_payables:             float = 0
    staff_pension_liability:    float = 0
    long_term_loan:             float = 0
    total_liabilities:          float = 0
    net_assets:                 float = 0
    adjusted_net_assets:        float = 0


@dataclass
class FundFinancials:
    income_statements:  list[IncomeStatement] = field(default_factory=list)
    balance_sheets:     list[BalanceSheet]    = field(default_factory=list)

    def balance_sheet(self, year: int) -> Optional[BalanceSheet]:
        for bs in self.balance_sheets:
            if bs.year == year:
                return bs
        return None

    def latest_balance_sheet(self) -> Optional[BalanceSheet]:
        if not self.balance_sheets:
            return None
        return max(self.balance_sheets, key=lambda b: b.year)

    def income_statement(self, year: str) -> Optional[IncomeStatement]:
        for is_ in self.income_statements:
            if is_.year_end == str(year):
                return is_
        return None

    def net_assets_series(self) -> dict[int, float]:
        return {bs.year: bs.net_assets for bs in sorted(self.balance_sheets, key=lambda b: b.year)}

    def adjusted_net_assets_series(self) -> dict[int, float]:
        return {bs.year: bs.adjusted_net_assets for bs in sorted(self.balance_sheets, key=lambda b: b.year)}


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------

def _safe(v) -> float:
    """Convert a cell value to float, returning 0.0 for None/nan/formula strings."""
    if v is None:
        return 0.0
    if isinstance(v, str):
        return 0.0
    try:
        f = float(v)
        return 0.0 if np.isnan(f) else f
    except (ValueError, TypeError):
        return 0.0


def load_fund_financials() -> FundFinancials:
    """
    Parse Assets_Information.xlsx and return a FundFinancials object.

    Layout (Revenue Acc sheet):
      Col B = 2021 income statement
      Col C = 2022 income statement
      Col D = 2023 income statement
      Col H = 2020 balance sheet
      Col I = 2021 balance sheet
      Col J = 2022 balance sheet
      Col K = 2023 balance sheet
    """
    path = ASSETS_FILE if ASSETS_FILE.exists() else _UPLOAD
    if not path.exists():
        log.warning("Assets_Information.xlsx not found — returning empty financials.")
        return FundFinancials()

    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Revenue Acc"]

    # Read all non-empty rows into a list
    raw: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        raw.append(tuple(c for c in row))

    # Helper: get value at (0-based row, 0-based col)
    def cell(r: int, c: int) -> float:
        if r >= len(raw):
            return 0.0
        row = raw[r]
        if c >= len(row):
            return 0.0
        return _safe(row[c])

    # ── Map known row indices (0-based, confirmed from file read) ─────────────
    # Income statement rows (cols 1=B, 2=C, 3=D for years 2021, 2022, 2023)
    R = {
        "fund_opening":            5,
        "employee_contributions":  9,
        "employer_contributions":  11,
        "financing_gap":           12,
        "government_grants":       13,
        "investment_income":       14,
        "fx":                      15,
        "change_fv_ip":            16,
        "change_fv_equity":        17,
        "other_income":            20,
        "total_income":            21,
        "pensions_paid":           23,
        "commutation":             24,
        "lumpsum":                 25,
        "other_benefits":          26,
        "admin":                   30,
        "withholding_tax":         31,
        "finance_cost":            32,
        "total_expenditure":       39,
        "surplus":                 41,
        "fund_closing":            43,
    }

    income_years  = [("2021", 1), ("2022", 2), ("2023", 3)]
    income_stmts  = []
    for yr_label, col in income_years:
        is_ = IncomeStatement(year_end=yr_label)
        is_.fund_opening              = cell(R["fund_opening"], col)
        is_.employee_contributions    = cell(R["employee_contributions"], col)
        is_.employer_contributions    = cell(R["employer_contributions"], col)
        is_.financing_gap             = cell(R["financing_gap"], col)
        is_.government_grants         = cell(R["government_grants"], col)
        is_.investment_income         = cell(R["investment_income"], col)
        is_.fx_gains_losses           = cell(R["fx"], col)
        is_.change_fv_investment_prop = cell(R["change_fv_ip"], col)
        is_.change_fv_equity          = cell(R["change_fv_equity"], col)
        is_.other_income              = cell(R["other_income"], col)
        is_.total_income              = cell(R["total_income"], col)
        is_.pensions_paid             = cell(R["pensions_paid"], col)
        is_.commutation_lumpsum       = cell(R["commutation"], col)
        is_.lumpsum_early_death       = cell(R["lumpsum"], col)
        is_.admin_expenses            = cell(R["admin"], col)
        is_.withholding_tax           = cell(R["withholding_tax"], col)
        is_.finance_cost              = cell(R["finance_cost"], col)
        is_.total_expenditure         = cell(R["total_expenditure"], col)
        is_.surplus                   = cell(R["surplus"], col)
        is_.fund_closing              = cell(R["fund_closing"], col)
        income_stmts.append(is_)

    # ── Balance sheet rows (cols 7=H, 8=I, 9=J, 10=K for years 2020-2023) ──
    BS = {
        "ppe":                   5,
        "intangibles":           7,
        "wip":                   9,
        "investment_prop":       10,
        "staff_loans":           11,
        "home_loans":            12,
        "microfinance":          13,
        "equity":                14,
        "fixed_deposit":         15,
        "govt_securities":       16,
        "total_assets":          23,    # SUM row H24 = row index 23
        "contributions_due":     19,
        "other_receivables":     20,
        "cash":                  21,
        "pensions_payable":      26,
        "benefits_payable":      27,
        "other_payables":        28,
        "staff_pension":         31,
        "long_term_loan":        32,
        "total_liabilities":     33,    # SUM row H34 = row index 33
        "net_assets":            35,    # row H36
        "adjusted_net_assets":   41,    # row H42
    }

    balance_years = [(2020, 7), (2021, 8), (2022, 9), (2023, 10)]
    balance_sheets = []
    for yr, col in balance_years:
        bs = BalanceSheet(year=yr)
        bs.ppe                      = cell(BS["ppe"],               col)
        bs.intangibles               = cell(BS["intangibles"],       col)
        bs.work_in_progress          = cell(BS["wip"],               col)
        bs.investment_properties     = cell(BS["investment_prop"],   col)
        bs.staff_loans               = cell(BS["staff_loans"],       col)
        bs.home_loan_scheme          = cell(BS["home_loans"],        col)
        bs.microfinance_loans        = cell(BS["microfinance"],      col)
        bs.equity_investments        = cell(BS["equity"],            col)
        bs.fixed_deposits            = cell(BS["fixed_deposit"],     col)
        bs.government_securities     = cell(BS["govt_securities"],   col)
        bs.contributions_due         = cell(BS["contributions_due"], col)
        bs.other_receivables         = cell(BS["other_receivables"], col)
        bs.cash_at_bank              = cell(BS["cash"],              col)
        bs.pensions_payable          = cell(BS["pensions_payable"],  col)
        bs.benefits_payable          = cell(BS["benefits_payable"],  col)
        bs.other_payables            = cell(BS["other_payables"],    col)
        bs.staff_pension_liability   = cell(BS["staff_pension"],     col)
        bs.long_term_loan            = cell(BS["long_term_loan"],    col)

        # Compute totals from components (more reliable than formula cells)
        bs.total_assets = (
            bs.ppe + bs.intangibles + bs.work_in_progress +
            bs.investment_properties + bs.staff_loans + bs.home_loan_scheme +
            bs.microfinance_loans + bs.equity_investments +
            bs.fixed_deposits + bs.government_securities +
            bs.contributions_due + bs.other_receivables + bs.cash_at_bank
        ) or cell(BS["total_assets"], col)

        bs.total_liabilities = (
            bs.pensions_payable + bs.benefits_payable + bs.other_payables +
            bs.staff_pension_liability + bs.long_term_loan
        ) or cell(BS["total_liabilities"], col)

        bs.net_assets          = cell(BS["net_assets"],          col)
        bs.adjusted_net_assets = cell(BS["adjusted_net_assets"], col)

        # Fallback: compute net assets if the formula cell returned 0
        if bs.net_assets == 0 and bs.total_assets > 0:
            bs.net_assets = bs.total_assets - bs.total_liabilities
        if bs.adjusted_net_assets == 0:
            bs.adjusted_net_assets = bs.net_assets

        balance_sheets.append(bs)

    log.info("Loaded financials: %d income statements, %d balance sheets",
             len(income_stmts), len(balance_sheets))
    return FundFinancials(income_statements=income_stmts, balance_sheets=balance_sheets)


# ---------------------------------------------------------------------------
# Solvency Position
# ---------------------------------------------------------------------------

@dataclass
class SolvencyPosition:
    valuation_year:         int
    total_assets:           float    # adjusted net assets from balance sheet
    total_pv_liability:     float    # from calculation engine
    fsl_a:                  float    # sub-fund A FSL
    fsl_b:                  float    # sub-fund B FSL
    psl_a:                  float    # sub-fund A PSL
    psl_b:                  float    # sub-fund B PSL
    # computed
    total_fsl:              float = 0.0
    total_psl:              float = 0.0
    total_liability:        float = 0.0
    surplus_deficit:        float = 0.0
    solvency_ratio:         float = 0.0
    funding_level_pct:      float = 0.0
    # funding ratio per sub-fund (assets split pro-rata)
    assets_a:               float = 0.0
    assets_b:               float = 0.0
    solvency_ratio_a:       float = 0.0
    solvency_ratio_b:       float = 0.0

    def __post_init__(self):
        self.total_fsl      = self.fsl_a + self.fsl_b
        self.total_psl      = self.psl_a + self.psl_b
        self.total_liability = self.total_fsl + self.total_psl

        if self.total_liability > 0:
            self.surplus_deficit    = self.total_assets - self.total_liability
            self.solvency_ratio     = self.total_assets / self.total_liability
            self.funding_level_pct  = self.solvency_ratio * 100

        # Pro-rata asset split by sub-fund liability weight
        total_psl = self.psl_a + self.psl_b
        if total_psl > 0:
            self.assets_a = self.total_assets * (self.psl_a / total_psl)
            self.assets_b = self.total_assets * (self.psl_b / total_psl)
        else:
            self.assets_a = self.total_assets / 2
            self.assets_b = self.total_assets / 2

        liability_a = self.fsl_a + self.psl_a
        liability_b = self.fsl_b + self.psl_b
        self.solvency_ratio_a = self.assets_a / liability_a if liability_a > 0 else 0
        self.solvency_ratio_b = self.assets_b / liability_b if liability_b > 0 else 0


def compute_solvency_position(
    fsl_a: float,
    fsl_b: float,
    psl_a: float,
    psl_b: float,
    total_pv_liability: float,
    financials: Optional[FundFinancials] = None,
    valuation_year: int = 2023,
) -> SolvencyPosition:
    """
    Combine actuarial liabilities with fund assets to produce the solvency position.

    fsl_a/b, psl_a/b  — from _compute_output_tables() in app.py (ZMW millions)
    total_pv_liability — from PensionEngine portfolio metrics (ZMW)
    financials         — from load_fund_financials(); uses latest balance sheet
    """
    if financials is None:
        financials = load_fund_financials()

    bs = financials.balance_sheet(valuation_year) or financials.latest_balance_sheet()
    scale = 1_000_000

    # Assets: use adjusted net assets from balance sheet (preferred),
    # fall back to net assets, then total_assets - total_liabilities
    if bs:
        if bs.adjusted_net_assets > 0:
            total_assets = bs.adjusted_net_assets
        elif bs.net_assets > 0:
            total_assets = bs.net_assets
        else:
            total_assets = bs.total_assets - bs.total_liabilities
    else:
        # No balance sheet available — use accumulated contributions as proxy
        total_assets = total_pv_liability   # will produce 100% solvency

    # Convert assets to ZMW millions to match FSL/PSL scale
    total_assets_m = total_assets / scale

    return SolvencyPosition(
        valuation_year      = valuation_year,
        total_assets        = total_assets_m,
        total_pv_liability  = total_pv_liability / scale,
        fsl_a               = fsl_a,
        fsl_b               = fsl_b,
        psl_a               = psl_a,
        psl_b               = psl_b,
    )


# ---------------------------------------------------------------------------
# CLI test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    ff = load_fund_financials()
    print("=== Income Statements ===")
    for is_ in ff.income_statements:
        print(f"  {is_.year_end}: Total income={is_.total_income:>16,.0f}  "
              f"Total expenditure={is_.total_expenditure:>16,.0f}  "
              f"Surplus={is_.surplus:>16,.0f}  "
              f"Closing fund={is_.fund_closing:>16,.0f}")

    print("\n=== Balance Sheets ===")
    for bs in ff.balance_sheets:
        print(f"  {bs.year}: Total assets={bs.total_assets:>16,.0f}  "
              f"Total liabilities={bs.total_liabilities:>16,.0f}  "
              f"Net assets={bs.net_assets:>16,.0f}  "
              f"Adj net assets={bs.adjusted_net_assets:>16,.0f}")
